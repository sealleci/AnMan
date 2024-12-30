from argparse import ArgumentParser
from dataclasses import dataclass
from os import chdir, listdir
from re import match
from subprocess import CalledProcessError, run
from sys import exit as exit_sys

from openpyxl import load_workbook


@dataclass
class PushArgs:
    message: str


def select_latest_xlsx_file(directory: str) -> str | None:
    TIANGAN_ORDER: list[str] = ["甲", "乙", "丙", "丁", "戊", "己", "庚", "辛", "壬", "癸"]
    filename_schema: str = r"^滿徽源集_版(.+)\.xlsx$"
    candidate_files: list[tuple[str, str]] = []

    for filename in listdir(directory):
        match_result = match(filename_schema, filename)

        if match_result:
            cur_tiangan = match_result.group(1)

            if not isinstance(cur_tiangan, str) or cur_tiangan not in TIANGAN_ORDER:
                continue

            candidate_files.append((filename, cur_tiangan))

    if len(candidate_files) == 0:
        return None

    candidate_files.sort(key=lambda x: TIANGAN_ORDER.index(x[1]), reverse=True)
    print(f"Selected xlsx file: {candidate_files[0][0]}")

    return f"{directory}/{candidate_files[0][0]}"


def read_names_from_xlsx(file_path: str) -> list[str]:
    workbook = load_workbook(file_path)
    cur_sheet = workbook["安滿與擬存在"]

    names: list[str] = []
    cur_row: int = 2

    while True:
        cell_value = cur_sheet.cell(row=cur_row, column=2).value

        if not cell_value:
            break

        if not isinstance(cell_value, str):
            continue

        names.append(cell_value)
        cur_row += 1

    return names


def write_data_to_file(data: list[str], output_file_path: str):
    with open(output_file_path, "w", encoding="utf-8") as f:
        for index, line in enumerate(data):
            f.write(f"{line}{"\n" if index != len(data) - 1 else ""}")

    print(f"Finished writing data to {output_file_path}")


def run_script(directory: str, script_filename: str):
    chdir(directory)

    try:
        run(["python", script_filename], check=True, capture_output=True, text=True)
        print(f"Executed {directory}/{script_filename} successfully")
    except CalledProcessError as e:
        print(f"Error occurred while executing {script_filename}: {e.stderr}")

    chdir("..")


def run_statistic():
    xlsx_file = select_latest_xlsx_file("./docs")

    if not xlsx_file:
        print("No valid xlsx file found")
        return

    write_data_to_file(read_names_from_xlsx(xlsx_file), "./amstat/input/chara.txt")
    run_script("./amstat", "count_chara.py")


def run_git_command(command_parts: list[str]):
    try:
        result = run(command_parts, check=True, text=True, capture_output=True)

        if result.stdout != "":
            print(result.stdout)
    except CalledProcessError as e:
        print(f"Error occurred while running {' '.join(command_parts)}: {e.stderr}")
        exit_sys(1)


def parse_args() -> PushArgs:
    parser = ArgumentParser(description="Git automation script")
    parser.add_argument("-m", "--message", required=True, help="Commit message")

    return PushArgs(**vars(parser.parse_args()))


def main(commit_message: str):
    print("> Running statistic script")
    run_statistic()

    print("> Adding changes")
    run_git_command(["git", "add", "."])

    print(f"> Committing changes with message: {commit_message}")
    run_git_command(["git", "commit", "-m", commit_message])

    print("> Pushing changes")
    run_git_command(["git", "push"])


if __name__ == "__main__":
    main(parse_args().message)
